"""
process_and_report.py
-----------------------
Reads the accumulated exchange rate history, cleans/reshapes it, and
produces:
 - Trend charts (PNG) per currency
 - A summary Excel workbook with a pivot table and live formulas

Run: python process_and_report.py
Input:  ./exchange_rate_history.csv
Output: ./*.png, ./Exchange_Rate_Summary.xlsx
"""

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

COLORS = {"ZAR": "#2E5A88", "EUR": "#4C8BF5", "GBP": "#7FB2F0", "JPY": "#F2A65A", "AUD": "#E85D75"}

df = pd.read_csv("./exchange_rate_history.csv", parse_dates=["rate_date"])

# --- Clean: drop exact duplicate (date, currency) pairs, keep latest fetch ---
df = df.sort_values("fetched_at").drop_duplicates(subset=["rate_date", "target_currency"], keep="last")
df = df.sort_values("rate_date").reset_index(drop=True)

# ---------- Chart: trend line per currency ----------
fig, ax = plt.subplots(figsize=(10, 6))
for currency, group in df.groupby("target_currency"):
    ax.plot(group["rate_date"], group["rate"], label=currency,
            color=COLORS.get(currency, "#888"), linewidth=2)
ax.set_title("USD Exchange Rate Trend (45-day history)", fontsize=14, fontweight="bold")
ax.set_ylabel("Rate (relative to 1 USD)")
ax.legend()
plt.xticks(rotation=30, ha="right")
plt.tight_layout()
plt.savefig("./exchange_rate_trend.png", dpi=150)
plt.close()

# ---------- Chart: latest rates snapshot ----------
latest_date = df["rate_date"].max()
latest = df[df["rate_date"] == latest_date].sort_values("rate", ascending=False)

fig, ax = plt.subplots(figsize=(7, 5))
bars = ax.bar(latest["target_currency"], latest["rate"],
              color=[COLORS.get(c, "#888") for c in latest["target_currency"]])
ax.set_title(f"Latest Rates vs USD ({latest_date.date()})", fontsize=14, fontweight="bold")
for b in bars:
    ax.text(b.get_x() + b.get_width()/2, b.get_height(), f"{b.get_height():.3f}",
            ha="center", va="bottom", fontsize=9)
plt.tight_layout()
plt.savefig("./latest_rates_snapshot.png", dpi=150)
plt.close()

print("Charts saved to ./")

# =========================================================
# Excel workbook
# =========================================================
wb = Workbook()
HEADER_FILL = PatternFill("solid", fgColor="2E5A88")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=13, color="2E5A88")


def style_header(ws, row=1, cols=None):
    cols = cols or ws.max_column
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autosize(ws):
    for col in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = length + 3


# ---- Sheet: Raw history ----
ws1 = wb.active
ws1.title = "Rate History"
ws1.append(["Date", "Currency", "Rate (per 1 USD)"])
for _, r in df.iterrows():
    ws1.append([r["rate_date"].strftime("%Y-%m-%d"), r["target_currency"], round(r["rate"], 4)])
style_header(ws1)
autosize(ws1)

# ---- Sheet: Pivot (date x currency) ----
ws2 = wb.create_sheet("Pivot - Rates by Date")
pivot = df.pivot_table(values="rate", index="rate_date", columns="target_currency")
ws2.append(["Date"] + list(pivot.columns))
for idx, row in pivot.iterrows():
    ws2.append([idx.strftime("%Y-%m-%d")] + [round(v, 4) for v in row])
style_header(ws2)
autosize(ws2)

line = LineChart()
line.title = "Exchange Rate Trend"
line.y_axis.title = "Rate"
line.x_axis.title = "Date"
data_ref = Reference(ws2, min_col=2, max_col=pivot.shape[1] + 1, min_row=1, max_row=pivot.shape[0] + 1)
cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=pivot.shape[0] + 1)
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cats_ref)
line.width, line.height = 24, 12
ws2.add_chart(line, f"A{pivot.shape[0] + 4}")

# ---- Sheet: KPI summary with live formulas ----
ws3 = wb.create_sheet("KPI Summary", 0)
ws3["A1"] = "Exchange Rate Tracker — Summary"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:B1")

n_rows = len(df)
ws3.append([])
ws3.append(["Metric", "Value"])
style_header(ws3, row=3)

kpis = [
    ("Tracking period (days of data)", f"=COUNTA('Pivot - Rates by Date'!A2:A{pivot.shape[0] + 1})"),
    ("Currencies tracked", len(df["target_currency"].unique())),
    ("Latest ZAR rate", f"=VLOOKUP(MAX('Pivot - Rates by Date'!A2:A{pivot.shape[0]+1}),'Pivot - Rates by Date'!A2:F{pivot.shape[0]+1},2,FALSE)"),
    ("Average ZAR rate (period)", f"=AVERAGE('Pivot - Rates by Date'!B2:B{pivot.shape[0]+1})"),
]
for label, value in kpis:
    ws3.append([label, value])
autosize(ws3)

wb.save("./Exchange_Rate_Summary.xlsx")
print("Excel workbook saved -> ./Exchange_Rate_Summary.xlsx")

print("\n--- Quick Summary ---")
print(f"Date range: {df['rate_date'].min().date()} to {df['rate_date'].max().date()}")
print(f"Currencies tracked: {', '.join(sorted(df['target_currency'].unique()))}")
print(f"Latest snapshot ({latest_date.date()}):")
for _, r in latest.iterrows():
    print(f"  1 USD = {r['rate']:.4f} {r['target_currency']}")
